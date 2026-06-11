import os
import json
import re
import io
import threading
import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from PIL import Image as PILImage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# =====================================================================
# 1. SERVICIO DE PROCESAMIENTO GRÁFICO EN MEMORIA (RAM)
# =====================================================================

def obtener_coordenadas_imagen(img):
    """Detecta de forma segura la fila y columna donde está anclada una foto."""
    try:
        if hasattr(img.anchor, '_from'):
            return img.anchor._from.row + 1, img.anchor._from.col + 1
        elif isinstance(img.anchor, str):
            match = re.match(r"([A-Z]+)(\d+)", img.anchor)
            if match:
                col_str, row_str = match.groups()
                row = int(row_str)
                col = sum((ord(char) - 64) * (26 ** i) for i, char in enumerate(reversed(col_str)))
                return row, col
    except Exception:
        pass
    return None, None

def extraer_imagen_proveedor_ram(ws_src, fila_excel):
    """Escanea la capa de dibujo buscando la foto anclada en la fila exacta."""
    for img in ws_src._images:
        row_idx, _ = obtener_coordenadas_imagen(img)
        if row_idx == fila_excel:
            return PILImage.open(img.ref)
    return None

def redimensionar_e_inyectar_imagen(pil_img, ws_dest, coordenada_celda, buffers_list):
    """
    Centra el producto vertical y horizontalmente dentro de un lienzo transparente 
    que ocupa el 100% de la celda, garantizando simetría y protegiendo las rejillas.
    """
    try:
        # 1. Convertir a RGBA para aislar canales de transparencia originales
        img_rgba = pil_img.convert("RGBA")
        
        # 2. AUTOCROP: Eliminar los márgenes transparentes vacíos del proveedor
        bbox = img_rgba.getbbox()
        if bbox:
            img_rgba = img_rgba.crop(bbox)

        # Rebana los 3 píxeles superiores sucios
        img_rgba = img_rgba.crop((0, 3, img_rgba.width, img_rgba.height))
        
        # 3. ELIMINAR LÍNEAS NEGRAS: Fusionar prenda sobre fondo blanco sólido antes del reescalado
        fondo_solido = PILImage.new("RGBA", img_rgba.size, (255, 255, 255, 255))
        fondo_solido.paste(img_rgba, (0, 0), img_rgba)
        img_limpia = fondo_solido.convert("RGB")
        
        # 4. ESCALAR PRENDA: Redimensionar el polo a un tamaño óptimo y estético
        max_width = 220
        max_height = 200
        img_limpia.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
        
        # 5. LIENZO TRANSPARENTE TOTAL (Ancho col 30 es ~210px. Alto de fila 170pt es ~226px reales)
        canvas_width = 210
        canvas_height = 226
        canvas = PILImage.new("RGBA", (canvas_width, canvas_height), (255, 255, 255, 0))
        
        # 6. CENTRADO MATEMÁTICO ABSOLUTO (Con offset de aire para el texto inferior)
        offset_x = (canvas_width - img_limpia.width) // 2
        offset_y = ((canvas_height - img_limpia.height) // 2) - 12
        if offset_y < 0:
            offset_y = 0
            
        canvas.paste(img_limpia, (offset_x, offset_y))
        
        # 7. Persistencia binaria en RAM (Formato PNG transparente)
        img_buffer = io.BytesIO()
        canvas.save(img_buffer, format="PNG")
        img_buffer.seek(0)
        
        buffers_list.append(img_buffer)
        
        # 8. INYECTAR EN EXCEL
        img_excel = OpenpyxlImage(img_buffer)
        img_excel.anchor = coordenada_celda
        
        ws_dest.add_image(img_excel)
        return True
    except Exception:
        return False

# =====================================================================
# 2. MOTOR DE CÁLCULO E INTERPOLACIÓN (BACKEND)
# =====================================================================

def calcular_margen_interpolado(cantidad, margenes_dict):
    """Calcula el margen exacto usando interpolación lineal."""
    cantidades = sorted([int(k) for k in margenes_dict.keys()])
    
    if cantidad <= cantidades[0]:
        return margenes_dict[str(cantidades[0])]
    if cantidad >= cantidades[-1]:
        return margenes_dict[str(cantidades[-1])]
        
    for i in range(len(cantidades) - 1):
        if cantidades[i] <= cantidad <= cantidades[i+1]:
            x1, x2 = cantidades[i], cantidades[i+1]
            y1, y2 = margenes_dict[str(x1)], margenes_dict[str(x2)]
            pendiente = (y2 - y1) / (x2 - x1)
            return round(y1 + pendiente * (cantidad - x1), 2)
    return 0.0

def limpiar_precio_proveedor(costo_str):
    if not costo_str:
        return 0.0
    num_match = re.search(r'[\d.]+', costo_str.replace(',', ''))
    return float(num_match.group()) if num_match else 0.0

def obtener_costo_proveedor_escalon(cantidad_target, celda_cantidades, celda_costos):
    lista_cants = [int(c.strip()) for c in str(celda_cantidades).split('\n') if c.strip().isdigit()]
    lista_costos_raw = [c.strip() for c in str(celda_costos).split('\n') if c.strip()]
    
    es_incluye_igv = any('incluye' in c.lower() for c in lista_costos_raw)
    lista_costos_clean = [c for c in lista_costos_raw if 'igv' not in c.lower()]
    
    if not lista_cants or not lista_costos_clean:
        return 0.0, False
        
    pares = []
    for i in range(len(lista_cants)):
        idx_costo = min(i, len(lista_costos_clean) - 1)
        pares.append((lista_cants[i], limpiar_precio_proveedor(lista_costos_clean[idx_costo])))
    
    pares = sorted(pares, key=lambda x: x[0])
    
    costo_proveedor = pares[0][1]
    for cant, costo in pares:
        if cant <= cantidad_target:
            costo_proveedor = costo
        else:
            break
            
    if es_incluye_igv:
        costo_proveedor = round(costo_proveedor / 1.18, 2)
        
    return costo_proveedor, es_incluye_igv

def procesar_cotizacion(producto_key, cantidad, config_json, path_proveedores):
    """Busca eficientemente en el Excel de proveedores, calcula y maqueta con estilos e imágenes."""
    prod_config = config_json[producto_key]
    nombre_comercial = prod_config["nombre_comercial"]
    filtros_prenda = [f.lower() for f in prod_config["filtros_prenda"]]
    filtros_material = [f.lower() for f in prod_config["filtros_material"]]
    
    excel_file = pd.ExcelFile(path_proveedores)
    pestañas_validas = []
    
    for sheet_name in excel_file.sheet_names:
        sheet_lower = sheet_name.lower()
        if any(f in sheet_lower for f in filtros_prenda) or "pantalones" in sheet_lower or "camisas" in sheet_lower:
            pestañas_validas.append(sheet_name)
            
    if not pestañas_validas:
        pestañas_validas = excel_file.sheet_names

    wb_prov_openpyxl = load_workbook(path_proveedores, data_only=True)

    filas_encontradas = []
    orígenes_hojas = {}
    indices_pandas = {}
    
    for sheet_name in pestañas_validas:
        df = pd.read_excel(path_proveedores, sheet_name=sheet_name, header=2)
        df.columns = [re.sub(r'\s+', ' ', str(c)).strip().upper().replace('Ò', 'O').replace('À', 'A').replace('È', 'E') for c in df.columns]
        
        if 'PRODUCTO' not in df.columns:
            continue
            
        for idx_pnd, row in df.iterrows():
            prod_celda = str(row.get('PRODUCTO', '')).lower()
            mat_celda = str(row.get('DETALLE', '')).lower() + " " + str(row.get('INFORMACION ADICIONAL', '')).lower()
            
            if any(f in prod_celda for f in filtros_prenda) and any(f in mat_celda for f in filtros_material):
                filas_encontradas.append(row)
                row_id = len(filas_encontradas) - 1
                orígenes_hojas[row_id] = sheet_name
                indices_pandas[row_id] = idx_pnd
                
    if not filas_encontradas:
        raise Exception(f"No se encontraron registros de proveedores para {nombre_comercial}.")
        
    margen_cliente = calcular_margen_interpolado(cantidad, prod_config["margenes"])
    
    output_dir = os.path.join("data", "output")
    os.makedirs(output_dir, exist_ok=True)
    
    nombre_seguro = nombre_comercial.replace("/", "-")
    base_name = f"cotizacion {cantidad} {nombre_seguro}"
    filename = os.path.join(output_dir, f"{base_name}.xlsx")
    
    counter = 1
    while os.path.exists(filename):
        filename = os.path.join(output_dir, f"{base_name}_v{counter}.xlsx")
        counter += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"
    ws.views.sheetView[0].showGridLines = True

    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="000000")
    data_font = Font(name="Arial", size=10)
    foto_font = Font(name="Arial", size=9, bold=True, color="FF0000")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    foto_alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    
    thin_side = Side(style='thin', color='B0B0B0')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    currency_format = '"S/." #,##0.00'

    all_headers = [
        "N°", "Proveedor", "Producto", "Foto", "Cant.", 
        "Costo uni. NO IGV (S/.)", "Tiempo Entrega", "Detalle", "Costo TOTAL NO IGV (S/.)"
    ]
    
    header_row = 3
    ws.row_dimensions[header_row].height = 28
    
    for col_idx, h_text in enumerate(all_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h_text)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = yellow_fill
        cell.border = thin_border

    active_buffers = []

    current_row = 4
    for i, row_prov in enumerate(filas_encontradas):
        ws.row_dimensions[current_row].height = 170  
        
        costo_prov_no_igv, _ = obtener_costo_proveedor_escalon(
            cantidad, 
            row_prov.get('CANTIDAD', ''), 
            row_prov.get('COSTO TOTAL', '')
        )
        
        costo_uni_cliente = round(costo_prov_no_igv * (1 + (margen_cliente / 100)), 2)
        costo_total_cliente = round(costo_uni_cliente * cantidad, 2)
        
        proveedor_nombre = str(row_prov.get('PROVEEDOR NUEVO', 'Anónimo')).strip()
        
        val_tiempo = row_prov.get('TIEMPO DE ENTREGA')
        if pd.isna(val_tiempo) or str(val_tiempo).strip().lower() in ['nan', '']:
            tiempo_entrega = "A coordinar"
        else:
            tiempo_entrega = str(val_tiempo).strip()
            
        detalle_celda = str(row_prov.get('DETALLE', '')).strip()
        
        sheet_origen = orígenes_hojas[i]
        idx_pnd_original = indices_pandas[i]
        fila_fisica_excel_proveedores = idx_pnd_original + 4
        
        ws_src_openpyxl = wb_prov_openpyxl[sheet_origen]
        imagen_pil = extraer_imagen_proveedor_ram(ws_src_openpyxl, fila_fisica_excel_proveedores)
        
        row_values = [
            i + 1,
            proveedor_nombre,
            nombre_comercial,
            "\n\n\nImagen Referencial",
            cantidad,
            costo_uni_cliente,
            tiempo_entrega,
            detalle_celda,
            costo_total_cliente
        ]
        
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx == 4:
                cell.font = foto_font
                cell.alignment = foto_alignment
                if imagen_pil is not None:
                    redimensionar_e_inyectar_imagen(imagen_pil, ws, f"D{current_row}", active_buffers)
            elif col_idx == 8:
                cell.font = data_font
                cell.alignment = left_alignment
            else:
                cell.font = data_font
                cell.alignment = center_alignment
                
            if col_idx in [6, 9]:
                cell.number_format = currency_format
                
        current_row += 1

    for col_idx in range(1, len(all_headers) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 8:
            ws.column_dimensions[col_letter].width = 80
        elif col_idx == 4:
            ws.column_dimensions[col_letter].width = 30  
        elif col_idx in [2, 3, 6, 7, 9]:
            ws.column_dimensions[col_letter].width = 22
        else:
            ws.column_dimensions[col_letter].width = 12

    wb.save(filename)
    
    for buf in active_buffers:
        buf.close()
        
    return filename

# =====================================================================
# 3. INTERFAZ GRÁFICA DE USUARIO (GUI CON TKINTER)
# =====================================================================

class CotizadorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Cotizaciones Rápidas 2026")
        
        window_width = 480
        window_height = 300
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_coor = (screen_width // 2) - (window_width // 2)
        y_coor = (screen_height // 2) - (window_height // 2)
        self.root.geometry(f"{window_width}x{window_height}+{x_coor}+{y_coor}")
        self.root.resizable(False, False)

        self.path_json = "config/matrices_margen.json"
        self.path_excel_prov = "config/Proovedores (version 1) 2026.xlsx"
        self.config_data = self.cargar_configuracion()
        
        style = ttk.Style()
        style.theme_use('clam')
        
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="GENERADOR DE COTIZACIONES", font=("Arial", 14, "bold")).pack(pady=10)
        
        ttk.Label(main_frame, text="Seleccione el Producto:").pack(anchor=tk.W, pady=2)
        self.combo_productos = ttk.Combobox(main_frame, state="readonly", width=30)
        self.combo_productos.pack(anchor=tk.W, pady=5)
        
        if self.config_data:
            self.mapeo_productos = {v["nombre_comercial"]: k for k, v in self.config_data.items()}
            self.combo_productos['values'] = list(self.mapeo_productos.keys())
            self.combo_productos.current(0)
            
        ttk.Label(main_frame, text="Cantidad de Unidades a Cotizar:").pack(anchor=tk.W, pady=2)
        self.entry_cantidad = ttk.Entry(main_frame, width=20, justify="center")
        self.entry_cantidad.pack(anchor=tk.W, pady=5)
        self.entry_cantidad.insert(0, "100")
        
        self.lbl_status = ttk.Label(main_frame, text="", font=("Arial", 10, "italic"), foreground="blue")
        self.lbl_status.pack(pady=10)
        
        self.btn_generar = ttk.Button(main_frame, text="Generar Cotización", command=self.iniciar_hilo_procesamiento)
        self.btn_generar.pack(fill=tk.X, pady=5)

    def cargar_configuracion(self):
        try:
            if os.path.exists(self.path_json):
                with open(self.path_json, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                messagebox.showerror("Error de Configuración", f"No se encontró el archivo: {self.path_json}")
                return {}
        except Exception as e:
            messagebox.showerror("Error", f"Error al leer JSON: {str(e)}")
            return {}

    def iniciar_hilo_procesamiento(self):
        nombre_comercial = self.combo_productos.get()
        qty_raw = self.entry_cantidad.get()
        
        if not nombre_comercial or not qty_raw.isdigit():
            messagebox.showwarning("Datos Inválidos", "Por favor introduce una cantidad numérica válida.")
            return
            
        cantidad = int(qty_raw)
        if cantidad <= 0:
            messagebox.showwarning("Datos Inválidos", "La cantidad debe ser mayor a 0.")
            return
            
        producto_key = self.mapeo_productos[nombre_comercial]
        
        if not os.path.exists(self.path_excel_prov):
            messagebox.showerror("Archivo Faltante", f"No se encontró la base de datos de proveedores:\n'{self.path_excel_prov}'")
            return

        self.btn_generar.config(state="disabled")
        self.lbl_status.config(text="Generando... Por favor espere.", foreground="dark orange")
        
        t = threading.Thread(target=self.tarea_background, args=(producto_key, cantidad))
        t.daemon = True
        t.start()

    def tarea_background(self, producto_key, cantidad):
        try:
            archivo_creado = procesar_cotizacion(producto_key, cantidad, self.config_data, self.path_excel_prov)
            self.root.after(0, self.proceso_exitoso, archivo_creado)
        except Exception as e:
            self.root.after(0, self.proceso_fallido, str(e))

    def proceso_exitoso(self, nombre_archivo):
        self.btn_generar.config(state="normal")
        self.lbl_status.config(text="🟢 ¡Archivo generado con éxito!", foreground="green")
        messagebox.showinfo("Éxito total", f"El archivo se generó correctamente:\n\n{nombre_archivo}")

    def proceso_fallido(self, error_msg):
        self.btn_generar.config(state="normal")
        self.lbl_status.config(text="🔴 Error en el proceso.", foreground="red")
        messagebox.showerror("Fallo de Generación", f"Hubo un problema al procesar:\n\n{error_msg}")

if __name__ == "__main__":
    if not os.path.exists("config"):
        os.makedirs("config")
        
    root = tk.Tk()
    app = CotizadorApp(root)
    root.mainloop()