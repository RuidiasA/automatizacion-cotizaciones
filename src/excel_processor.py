import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook
from src.matcher import ProductMatcher
from src.calculator import QuoteCalculator


class ExcelProcessor:
    def __init__(self, json_config, input_dir="data/input", output_dir="data/output"):
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.matcher = ProductMatcher(json_config)
        self.calculator = QuoteCalculator(json_config)

        # Aseguramos que existan las carpetas de entrada y salida
        os.makedirs(self.input_dir, exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)

    def _normalize_header(self, header_name):
        if pd.isna(header_name) or not header_name:
            return ""
        text = str(header_name).lower().strip()
        # Quitamos tildes básicas para normalizar encabezados
        replacements = (("á", "a"), ("é", "e"),
                        ("í", "i"), ("ó", "o"), ("ú", "u"))
        for target, replacement in replacements:
            text = text.replace(target, replacement)
        return text

    def process_all_excels(self):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        excel_files = glob.glob(os.path.join(
            self.input_dir, "*.xlsx")) + glob.glob(os.path.join(self.input_dir, "*.xls"))
        if not excel_files:
            print(
                f"⚠️ No se encontraron archivos de Excel en '{self.input_dir}'.")
            return

        print(f"🚀 Se encontraron {len(excel_files)} archivos para procesar.")
        output_file_path = os.path.join(
            self.output_dir, "consolidado_cotizaciones.xlsx")

        from openpyxl import Workbook
        wb_output = Workbook()
        wb_output.remove(wb_output.active)

        used_sheet_names = []

        # Instancias de Estilos Únicas (Eficiencia de Memoria XML)
        yellow_fill = PatternFill(
            start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="000000")
        title_font = Font(name="Arial", size=14, bold=True, color="000000")
        data_font = Font(name="Arial", size=10)
        foto_font = Font(name="Arial", size=9, bold=True, color="FF0000")

        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True)

        title_alignment = Alignment(horizontal="center", vertical="center")
        foto_alignment = Alignment(
            horizontal="center", vertical="bottom", wrap_text=True)

        thin_side = Side(style='thin', color='B0B0B0')
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)
        currency_format = '"S/." #,##0.00'
        percentage_format = '0.00%'

        for file_path in excel_files:
            file_name = os.path.basename(file_path)

            if file_name.startswith("~$"):
                continue

            clean_name = os.path.splitext(file_name)[0]
            sheet_tab_name = clean_name[:26].strip()

            base_name = sheet_tab_name
            counter = 1
            while sheet_tab_name in used_sheet_names:
                sheet_tab_name = f"{base_name}_{counter}"
                counter += 1

            try:
                xls = pd.ExcelFile(file_path)
                all_processed_rows = []

                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(
                        file_path, sheet_name=sheet_name, header=None)
                    rows_calculated = self._process_sheet_data(
                        df, sheet_name, file_name)
                    if rows_calculated:
                        all_processed_rows.extend(rows_calculated)

                if not all_processed_rows:
                    continue

                used_sheet_names.append(sheet_tab_name)
                print(
                    f"🎨 Formateando y procesando: {file_name} -> Pestaña: {sheet_tab_name}")

                ws = wb_output.create_sheet(title=sheet_tab_name)
                ws.views.sheetView[0].showGridLines = True

                ws.merge_cells("A1:I1")
                ws["A1"] = clean_name.upper().strip()
                ws["A1"].font = title_font
                ws["A1"].alignment = title_alignment
                ws.row_dimensions[1].height = 25

                client_headers = ["N°", "Cód. Artíc.", "Art.", "Foto", "Cant.",
                                  "Costo Uni. NO IGV(S/.)", "Tiempo Entrega", "Detalle", "Costo TOTAL NO IGV (S/.)"]

                # Inyectamos tus nuevas cabeceras de control en el orden exacto solicitado
                control_headers = [
                    "Origen (Hoja)",
                    "[CTRL] Costo Uni. NO IGV(S/.)",
                    "[CTRL] COSTO PROVEEDOR",
                    "[CTRL] % VARIACIÓN",
                    "[CTRL] ESTADO VALIDACIÓN",
                    "[CTRL] % MARGEN APLICADO"
                ]
                all_headers = client_headers + control_headers

                header_row = 3
                ws.row_dimensions[header_row].height = 28

                # Renderizar Encabezados
                for col_idx, h_text in enumerate(all_headers, start=1):
                    cell = ws.cell(
                        row=header_row, column=col_idx, value=h_text)
                    cell.font = header_font
                    cell.alignment = center_alignment

                    if col_idx <= len(client_headers):
                        cell.fill = yellow_fill
                        cell.border = thin_border

                # Renderizar Filas de Productos
                current_row = 4
                for idx, data_row in enumerate(all_processed_rows, start=1):
                    ws.row_dimensions[current_row].height = 95

                    # 💡 FÓRMULA DE VARIACIÓN SOLICITADA:
                    # Aplica la variación entre la nueva columna K (Costo Uni extraído del excel original) y la columna L (Costo Proveedor)
                    # En coordenadas de Excel: Columna K es la 11 y Columna L es la 12
                    formula_variacion = f"=(K{current_row}/L{current_row})-1"

                    row_values = [
                        # A (1)
                        idx,
                        # B (2)
                        data_row["Cód. Artíc."],
                        # C (3)
                        data_row["Art."],
                        # D (4)
                        data_row["Foto"],
                        # E (5)
                        data_row["Cant."],
                        # F (6) -> Calculado con tu JSON
                        data_row["Costo Uni. NO IGV(S/.)"],
                        # G (7)
                        data_row["Tiempo Entrega"],
                        # H (8)
                        data_row["Detalle"],
                        # I (9)
                        data_row["Costo TOTAL NO IGV (S/.)"],
                        # J (10)
                        data_row["Origen (Hoja)"],
                        # K (11) -> El extraído crudo de tus 77 Excels
                        data_row["[CTRL] Costo Uni. Original"],
                        # L (12) -> Costo Proveedor
                        data_row["[CTRL] COSTO PROVEEDOR"],
                        # M (13) -> Variación K vs L
                        formula_variacion,
                        # N (14)
                        data_row["[CTRL] ESTADO VALIDACIÓN"],
                        # O (15)
                        data_row["[CTRL] % MARGEN APLICADO"],
                    ]

                    for col_idx, value in enumerate(row_values, start=1):
                        cell = ws.cell(row=current_row,
                                       column=col_idx, value=value)

                        if col_idx == 4:
                            cell.font = foto_font
                            cell.alignment = foto_alignment
                        elif col_idx == 8:
                            cell.font = data_font
                            cell.alignment = left_alignment
                        else:
                            cell.font = data_font
                            cell.alignment = center_alignment

                        # Formato numérico seguro de Moneda (Agregamos la columna 11 y 12)
                        if col_idx in [6, 9, 11, 12]:
                            cell.number_format = currency_format

                        # Formato para la columna de porcentaje de variación (Columna 13)
                        if col_idx == 13:
                            cell.number_format = percentage_format

                        if col_idx <= len(client_headers):
                            cell.border = thin_border

                    current_row += 1

                    if current_row % 100 == 0:
                        import gc
                        gc.collect()

                # Ajuste de Anchos de Columnas considerando las nuevas dimensiones
                for col_idx in range(1, len(all_headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_idx == 8:
                        ws.column_dimensions[col_letter].width = 80
                    elif col_idx in [3, 6, 7, 9, 11, 12, 15]:
                        ws.column_dimensions[col_letter].width = 22
                    elif col_idx == 4:
                        ws.column_dimensions[col_letter].width = 15
                    else:
                        ws.column_dimensions[col_letter].width = 12

            except Exception as e:
                print(
                    f"❌ Error de maquetación en el archivo {file_name}: {str(e)}")

        if len(wb_output.sheetnames) == 0:
            ws_empty = wb_output.create_sheet(title="Sin Coincidencias")
            ws_empty["A1"] = "Ningún producto de los excels coincidió con tus matrices JSON."

        wb_output.save(output_file_path)
        print(
            f"✨ ¡Automatización completada con éxito! Archivo limpio y maquetado en: {output_file_path}")

    def _process_sheet_data(self, df, sheet_name, file_name):
        rows_to_return = []

        col_detalle_idx = None
        col_cantidad_idx = None
        col_costo_cliente_idx = None
        col_costo_proveedor_idx = None

        header_row_idx = None
        for row_idx in range(min(len(df), 20)):
            row_values = [self._normalize_header(
                val) for val in df.iloc[row_idx]]

            if any(h in ["detalle", "detalles"] for h in row_values):
                header_row_idx = row_idx
                break

        if header_row_idx is None:
            return []

        headers = [self._normalize_header(val)
                   for val in df.iloc[header_row_idx]]

        col_tiempo_idx = None

        for idx, h in enumerate(headers):
            if h in ["detalle", "detalles"]:
                col_detalle_idx = idx
            elif h in ["cant.", "cant", "cantidad"]:
                col_cantidad_idx = idx
            elif h.startswith("costo uni"):
                col_costo_cliente_idx = idx
            elif h == "costo s/.":
                col_costo_proveedor_idx = idx

            if h in ["tiempo entrega", "tiempo de entrega", "entrega"]:
                col_tiempo_idx = idx

        if None in [col_detalle_idx, col_cantidad_idx, col_costo_cliente_idx, col_costo_proveedor_idx]:
            return []

        for row_idx in range(header_row_idx + 1, len(df)):
            row_data = df.iloc[row_idx]

            detail_content = str(row_data.iloc[col_detalle_idx])
            if pd.isna(row_data.iloc[col_detalle_idx]) or detail_content.strip() == "":
                continue

            matched_product_key = self.matcher.find_matched_product(
                detail_content)

            if matched_product_key:
                try:
                    quantity = int(
                        float(str(row_data.iloc[col_cantidad_idx]).replace(",", "")))
                    provider_cost = float(str(row_data.iloc[col_costo_proveedor_idx]).replace(
                        "S/.", "").replace("S/", "").strip())
                    tiempo_entrega = str(row_data.iloc[col_tiempo_idx]).strip(
                    ) if col_tiempo_idx is not None else "5 a 7 Días útiles"

                    # 💡 EXTRAEMOS EL VALOR ORIGINAL CRUDO:
                    # Jalamos el valor original directo del Excel de entrada sin aplicar matemáticas de tu JSON
                    raw_original_cost = float(str(row_data.iloc[col_costo_cliente_idx]).replace(
                        "S/.", "").replace("S/", "").replace(",", "").strip())

                    calculated_cost, margin_used, status = self.calculator.calculate_client_cost(
                        matched_product_key, quantity, provider_cost
                    )
                    costo_total_calculado = round(
                        quantity * calculated_cost, 2)

                    processed_row = {
                        "N°": None,
                        "Cód. Artíc.": "",
                        "Art.": matched_product_key,
                        "Foto": "\n\n\nImagen Referencial",
                        "Cant.": quantity,
                        "Costo Uni. NO IGV(S/.)": calculated_cost,
                        "Tiempo Entrega": tiempo_entrega,
                        "Detalle": detail_content,
                        "Costo TOTAL NO IGV (S/.)": costo_total_calculado,

                        "Origen (Hoja)": sheet_name,
                        # 💡 Aquí guardamos el valor crudo original jalarlo al reporte
                        "[CTRL] Costo Uni. Original": raw_original_cost,
                        "[CTRL] COSTO PROVEEDOR": provider_cost,
                        "[CTRL] % MARGEN APLICADO": f"{round(margin_used, 2)}%",
                        "[CTRL] ESTADO VALIDACIÓN": status
                    }
                    rows_to_return.append(processed_row)

                except Exception as ex:
                    continue

        return rows_to_return
